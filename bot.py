import asyncio
import logging
import os

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode, ChatMemberStatus
from aiogram.filters import CommandStart, ChatMemberUpdatedFilter, IS_NOT_MEMBER, MEMBER
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    ChatMemberUpdated,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
)
from dotenv import load_dotenv

import database as db

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_IDS = [int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip()]

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
logger.info(f"ADMIN_IDS: {ADMIN_IDS}")

router = Router()

# ── Button texts ──────────────────────────────────────────────

BTN_REFERRALS = "👥 Mening referallarim"
BTN_TOP = "🏆 Top 10"
BTN_PRIZES = "🎁 Sovg'alar"
BTN_ABOUT = "ℹ️ Biz haqimizda"
BTN_ADMIN = "⚙️ Admin panel"
BTN_BACK = "◀️ Ortga"

BTN_ADM_BROADCAST = "📢 Xabar yuborish"
BTN_ADM_PRIZES = "🎁 Sovg'alar matni"
BTN_ADM_ABOUT = "ℹ️ Biz haqimizda matni"
BTN_ADM_WELCOME = "👋 Xush kelibsiz matni"
BTN_ADM_CHANNELS = "📡 Kanallar"
BTN_ADM_STATS = "📊 Statistika"
BTN_ADM_USERS = "📋 Foydalanuvchilar"

BTN_CH_ADD = "➕ Kanal qo'shish"
BTN_CH_REMOVE = "➖ Kanal o'chirish"
BTN_CH_LIST = "📋 Kanallar ro'yxati"


# ── States ────────────────────────────────────────────────────


class Registration(StatesGroup):
    first_name = State()
    last_name = State()
    phone = State()


class AdminBroadcast(StatesGroup):
    waiting_message = State()


class AdminEditSetting(StatesGroup):
    waiting_text = State()


class AdminChannelAdd(StatesGroup):
    waiting_channel_id = State()
    waiting_channel_url = State()


class AdminChannelRemove(StatesGroup):
    waiting_channel_id = State()


# ── Keyboards ─────────────────────────────────────────────────


def main_menu_kb(user_id: int) -> ReplyKeyboardMarkup:
    buttons = [
        [KeyboardButton(text=BTN_REFERRALS), KeyboardButton(text=BTN_TOP)],
        [KeyboardButton(text=BTN_PRIZES), KeyboardButton(text=BTN_ABOUT)],
    ]
    if user_id in ADMIN_IDS:
        buttons.append([KeyboardButton(text=BTN_ADMIN)])
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)


def admin_menu_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=BTN_ADM_BROADCAST), KeyboardButton(text=BTN_ADM_STATS)],
            [KeyboardButton(text=BTN_ADM_PRIZES), KeyboardButton(text=BTN_ADM_ABOUT)],
            [KeyboardButton(text=BTN_ADM_WELCOME), KeyboardButton(text=BTN_ADM_CHANNELS)],
            [KeyboardButton(text=BTN_ADM_USERS)],
            [KeyboardButton(text=BTN_BACK)],
        ],
        resize_keyboard=True,
    )


def channels_menu_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=BTN_CH_ADD), KeyboardButton(text=BTN_CH_REMOVE)],
            [KeyboardButton(text=BTN_CH_LIST)],
            [KeyboardButton(text=BTN_BACK)],
        ],
        resize_keyboard=True,
    )


def back_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=BTN_BACK)]],
        resize_keyboard=True,
    )


# ── Helpers ───────────────────────────────────────────────────


async def check_all_channels(bot: Bot, user_id: int) -> list[dict]:
    """Check subscription to all mandatory channels. Returns list of unsubscribed channels."""
    channels = await db.get_all_channels()
    if not channels:
        return []

    unsubscribed = []
    for ch in channels:
        try:
            member = await bot.get_chat_member(chat_id=ch["channel_id"], user_id=user_id)
            if member.status not in (
                ChatMemberStatus.MEMBER,
                ChatMemberStatus.ADMINISTRATOR,
                ChatMemberStatus.CREATOR,
            ):
                unsubscribed.append(dict(ch))
        except Exception:
            pass  # Skip broken channels
    return unsubscribed


async def send_subscribe_message(message: Message, unsubscribed: list[dict]):
    """Show buttons for all unsubscribed channels."""
    buttons = []
    for ch in unsubscribed:
        title = ch.get("title") or ch["channel_id"]
        buttons.append(
            [InlineKeyboardButton(text=f"📢 {title}", url=ch["channel_url"])]
        )
    buttons.append(
        [InlineKeyboardButton(text="✅ Tekshirish", callback_data="check_sub")]
    )

    await message.answer(
        "⚠️ <b>Konkursda qatnashish uchun quyidagi kanallarga a'zo bo'lishingiz shart!</b>\n\n"
        "Kanallarga a'zo bo'ling, so'ng «Tekshirish» tugmasini bosing.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


async def is_channel_event(event: ChatMemberUpdated) -> bool:
    """Check if this event is from one of our mandatory channels."""
    channels = await db.get_all_channels()
    for ch in channels:
        if str(event.chat.id) == ch["channel_id"]:
            return True
        if event.chat.username and f"@{event.chat.username}" == ch["channel_id"]:
            return True
    return False


# ── Channel member tracking ───────────────────────────────────


@router.chat_member(ChatMemberUpdatedFilter(member_status_changed=MEMBER >> IS_NOT_MEMBER))
async def on_user_left_channel(event: ChatMemberUpdated):
    if not await is_channel_event(event):
        return

    user_id = event.new_chat_member.user.id
    user = await db.get_user(user_id)
    if not user or not user["is_registered"]:
        return

    await db.deactivate_user(user_id)

    if user["referred_by"]:
        try:
            ref_count = await db.get_referral_count(user["referred_by"])
            await event.bot.send_message(
                user["referred_by"],
                f"😔 <b>{user['first_name']} {user['last_name']}</b> kanaldan chiqdi "
                f"va referallaringizdan olib tashlandi.\n\n"
                f"📊 Hozirgi referallaringiz: <b>{ref_count}</b>",
            )
        except Exception:
            pass


@router.chat_member(ChatMemberUpdatedFilter(member_status_changed=IS_NOT_MEMBER >> MEMBER))
async def on_user_joined_channel(event: ChatMemberUpdated):
    if not await is_channel_event(event):
        return

    user_id = event.new_chat_member.user.id
    user = await db.get_user(user_id)
    if not user or not user["is_registered"]:
        return

    await db.activate_user(user_id)

    if user["referred_by"]:
        try:
            ref_count = await db.get_referral_count(user["referred_by"])
            await event.bot.send_message(
                user["referred_by"],
                f"🎉 <b>{user['first_name']} {user['last_name']}</b> kanalga qaytib qo'shildi!\n\n"
                f"📊 Hozirgi referallaringiz: <b>{ref_count}</b>",
            )
        except Exception:
            pass


# ── Check subscription callback ───────────────────────────────


@router.callback_query(F.data == "check_sub")
async def check_sub_callback(callback: CallbackQuery, state: FSMContext):
    unsub = await check_all_channels(callback.bot, callback.from_user.id)
    if unsub:
        await callback.answer("❌ Siz hali barcha kanallarga a'zo bo'lmadingiz!", show_alert=True)
        return

    await callback.message.edit_text("✅ Barcha kanallarga a'zo bo'ldingiz! Davom etamiz...")

    if await db.is_registered(callback.from_user.id):
        welcome = await db.get_setting("welcome_text")
        await callback.message.answer(
            welcome, reply_markup=main_menu_kb(callback.from_user.id)
        )
    else:
        await callback.message.answer(
            "📝 <b>Ro'yxatdan o'tish</b>\n\nIsmingizni kiriting:",
            reply_markup=ReplyKeyboardRemove(),
        )
        await state.set_state(Registration.first_name)


# ── /start ────────────────────────────────────────────────────


@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    user_id = message.from_user.id
    username = message.from_user.username or ""

    # Parse referral
    referrer_id = None
    args = message.text.split()
    if len(args) > 1 and args[1].startswith("ref_"):
        try:
            referrer_id = int(args[1][4:])
            if referrer_id == user_id:
                referrer_id = None
        except ValueError:
            referrer_id = None

    # Save user stub
    existing = await db.get_user(user_id)
    if not existing:
        await db.add_user(user_id, username, referrer_id)

    # Check channels
    unsub = await check_all_channels(message.bot, user_id)
    if unsub:
        await send_subscribe_message(message, unsub)
        return

    # Already registered
    if await db.is_registered(user_id):
        await db.activate_user(user_id)
        welcome = await db.get_setting("welcome_text")
        await message.answer(welcome, reply_markup=main_menu_kb(user_id))
        return

    # Start registration
    welcome = await db.get_setting("welcome_text")
    await message.answer(welcome)
    await message.answer(
        "📝 <b>Ro'yxatdan o'tish</b>\n\nIsmingizni kiriting:",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Registration.first_name)


# ── Registration ──────────────────────────────────────────────


@router.message(Registration.first_name)
async def reg_first_name(message: Message, state: FSMContext):
    if message.text == BTN_BACK:
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=main_menu_kb(message.from_user.id))
        return
    await state.update_data(first_name=message.text.strip())
    await message.answer("Familiyangizni kiriting:")
    await state.set_state(Registration.last_name)


@router.message(Registration.last_name)
async def reg_last_name(message: Message, state: FSMContext):
    if message.text == BTN_BACK:
        await message.answer("Ismingizni kiriting:")
        await state.set_state(Registration.first_name)
        return
    await state.update_data(last_name=message.text.strip())
    contact_kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📱 Raqamni yuborish", request_contact=True)],
            [KeyboardButton(text=BTN_BACK)],
        ],
        resize_keyboard=True,
        one_time_keyboard=True,
    )
    await message.answer(
        "Telefon raqamingizni yuboring:\n"
        "(Tugmani bosing yoki qo'lda kiriting: +998901234567)",
        reply_markup=contact_kb,
    )
    await state.set_state(Registration.phone)


@router.message(Registration.phone, F.contact)
async def reg_phone_contact(message: Message, state: FSMContext):
    phone = message.contact.phone_number
    await _finish_registration(message, state, phone)


@router.message(Registration.phone)
async def reg_phone_text(message: Message, state: FSMContext):
    if message.text == BTN_BACK:
        await message.answer("Familiyangizni kiriting:", reply_markup=ReplyKeyboardRemove())
        await state.set_state(Registration.last_name)
        return
    phone = message.text.strip()
    if not phone.replace("+", "").replace(" ", "").isdigit() or len(phone) < 9:
        await message.answer("❌ Noto'g'ri raqam. Qaytadan kiriting (+998901234567):")
        return
    await _finish_registration(message, state, phone)


async def _finish_registration(message: Message, state: FSMContext, phone: str):
    data = await state.get_data()
    user_id = message.from_user.id

    await db.update_user_registration(
        user_id=user_id,
        first_name=data["first_name"],
        last_name=data["last_name"],
        phone=phone,
    )
    await state.clear()

    # Notify referrer
    user = await db.get_user(user_id)
    if user and user["referred_by"]:
        try:
            ref_count = await db.get_referral_count(user["referred_by"])
            await message.bot.send_message(
                user["referred_by"],
                f"🎉 Yangi referal ro'yxatdan o'tdi!\n"
                f"👤 {data['first_name']} {data['last_name']}\n"
                f"📊 Jami referallaringiz: <b>{ref_count}</b>",
            )
        except Exception:
            pass

    await message.answer(
        f"✅ <b>Tabriklaymiz, {data['first_name']}!</b>\n\n"
        "Siz muvaffaqiyatli ro'yxatdan o'tdingiz! 🎉\n"
        "Do'stlaringizni taklif qilib, sovg'alar yutib oling!",
        reply_markup=main_menu_kb(user_id),
    )


# ── Main menu ─────────────────────────────────────────────────


@router.message(F.text == BTN_REFERRALS)
async def my_referrals(message: Message):
    if not await db.is_registered(message.from_user.id):
        await message.answer("Avval ro'yxatdan o'ting: /start")
        return
    unsub = await check_all_channels(message.bot, message.from_user.id)
    if unsub:
        await send_subscribe_message(message, unsub)
        return

    user_id = message.from_user.id
    bot_info = await message.bot.get_me()
    ref_link = f"https://t.me/{bot_info.username}?start=ref_{user_id}"
    ref_count = await db.get_referral_count(user_id)
    referrals = await db.get_referrals(user_id)

    text = (
        f"👥 <b>Sizning referallaringiz</b>\n\n"
        f"📊 Faol referallar: <b>{ref_count}</b> nafar\n\n"
        f"🔗 <b>Referal havolangiz:</b>\n"
        f"<code>{ref_link}</code>\n\n"
        "☝️ Havolani do'stlaringizga yuboring!\n"
    )

    if referrals:
        text += "\n📋 <b>Referallar:</b>\n"
        for i, ref in enumerate(referrals[:15], 1):
            status = "✅" if ref["is_active"] else "❌"
            text += f"  {i}. {status} {ref['first_name']} {ref['last_name']}\n"
        text += "\n✅ faol  ❌ kanaldan chiqgan"

    share_kb = InlineKeyboardMarkup(
        inline_keyboard=[[
            InlineKeyboardButton(
                text="📤 Do'stlarga yuborish",
                url=f"https://t.me/share/url?url={ref_link}&text=PolyglotLC konkursida ishtirok et va sovg'alar yutib ol! 🎁",
            )
        ]]
    )
    await message.answer(text, reply_markup=share_kb)


@router.message(F.text == BTN_TOP)
async def top_10(message: Message):
    if not await db.is_registered(message.from_user.id):
        await message.answer("Avval ro'yxatdan o'ting: /start")
        return
    unsub = await check_all_channels(message.bot, message.from_user.id)
    if unsub:
        await send_subscribe_message(message, unsub)
        return

    top = await db.get_top_referrers(10)
    if not top:
        await message.answer("🏆 Hozircha hech kim referal taklif qilmagan.\n\nBirinchi bo'ling! 🚀")
        return

    medals = {1: "🥇", 2: "🥈", 3: "🥉"}
    text = "🏆 <b>Top 10 — eng ko'p referal taklif qilganlar</b>\n\n"
    for i, user in enumerate(top, 1):
        medal = medals.get(i, f"  {i}.")
        name = f"{user['first_name']} {user['last_name']}"
        text += f"{medal} <b>{name}</b> — {user['ref_count']} ta referal\n"

    my_count = await db.get_referral_count(message.from_user.id)
    text += f"\n📊 Sizning referallaringiz: <b>{my_count}</b>"
    await message.answer(text)


@router.message(F.text == BTN_PRIZES)
async def prizes(message: Message):
    text = await db.get_setting("prizes_text")
    if text:
        await message.answer(text)
    else:
        await message.answer("🎁 Sovg'alar haqida ma'lumot hozircha mavjud emas.")


@router.message(F.text == BTN_ABOUT)
async def about_us(message: Message):
    text = await db.get_setting("about_text")
    if text:
        await message.answer(text)
    else:
        await message.answer("ℹ️ Ma'lumot hozircha kiritilmagan.")


# ══════════════════════════════════════════════════════════════
#  ADMIN PANEL
# ══════════════════════════════════════════════════════════════


@router.message(F.text == BTN_ADMIN)
async def admin_panel(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    await state.clear()

    user_count = await db.get_user_count()
    active_count = await db.get_active_user_count()
    channels = await db.get_all_channels()

    await message.answer(
        f"🔧 <b>Admin panel</b>\n\n"
        f"👥 Ro'yxatdan o'tganlar: <b>{user_count}</b>\n"
        f"✅ Faol: <b>{active_count}</b>\n"
        f"📡 Kanallar soni: <b>{len(channels)}</b>",
        reply_markup=admin_menu_kb(),
    )


# ── Admin: Back ───────────────────────────────────────────────


@router.message(F.text == BTN_BACK)
async def go_back(message: Message, state: FSMContext):
    current = await state.get_state()
    await state.clear()

    if message.from_user.id in ADMIN_IDS and current and "Admin" in current:
        # From admin sub-menus → admin panel
        await admin_panel(message, state)
    else:
        # To main menu
        await message.answer("🏠 Bosh menyu", reply_markup=main_menu_kb(message.from_user.id))


# ── Admin: Broadcast ──────────────────────────────────────────


@router.message(F.text == BTN_ADM_BROADCAST)
async def admin_broadcast_start(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    await message.answer(
        "📢 <b>Xabar yuborish</b>\n\n"
        "Barcha foydalanuvchilarga yuboriladigan xabarni yozing.\n"
        "Matn, rasm, video — istalganini yuboring.",
        reply_markup=back_kb(),
    )
    await state.set_state(AdminBroadcast.waiting_message)


@router.message(AdminBroadcast.waiting_message)
async def admin_broadcast_send(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    if message.text == BTN_BACK:
        await state.clear()
        from_state = state
        await admin_panel(message, from_state)
        return

    await state.clear()
    user_ids = await db.get_all_user_ids()
    sent = 0
    failed = 0

    status_msg = await message.answer(f"📤 Yuborilmoqda... 0/{len(user_ids)}")

    for i, uid in enumerate(user_ids):
        try:
            await message.copy_to(uid)
            sent += 1
        except Exception:
            failed += 1
        if (i + 1) % 25 == 0:
            try:
                await status_msg.edit_text(f"📤 Yuborilmoqda... {i + 1}/{len(user_ids)}")
            except Exception:
                pass
        await asyncio.sleep(0.05)

    await status_msg.edit_text(
        f"✅ <b>Xabar yuborildi!</b>\n\n"
        f"📨 Yuborildi: {sent}\n"
        f"❌ Yuborilmadi: {failed}\n"
        f"👥 Jami: {len(user_ids)}"
    )
    await message.answer("🔧 Admin panel", reply_markup=admin_menu_kb())


# ── Admin: Edit settings ─────────────────────────────────────


EDIT_BUTTONS = {
    BTN_ADM_PRIZES: ("prizes_text", "🎁 Sovg'alar"),
    BTN_ADM_ABOUT: ("about_text", "ℹ️ Biz haqimizda"),
    BTN_ADM_WELCOME: ("welcome_text", "👋 Xush kelibsiz"),
}


@router.message(F.text.in_(EDIT_BUTTONS.keys()))
async def admin_edit_start(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return

    key, label = EDIT_BUTTONS[message.text]
    current = await db.get_setting(key) or "(bo'sh)"

    await state.update_data(setting_key=key, setting_label=label)
    await message.answer(
        f"✏️ <b>{label} matnini o'zgartirish</b>\n\n"
        f"📄 Hozirgi matn:\n\n{current}\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Yangi matnni yozing (HTML format):",
        reply_markup=back_kb(),
    )
    await state.set_state(AdminEditSetting.waiting_text)


@router.message(AdminEditSetting.waiting_text)
async def admin_edit_save(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    if message.text == BTN_BACK:
        await state.clear()
        await admin_panel(message, state)
        return

    data = await state.get_data()
    await db.set_setting(data["setting_key"], message.text)
    await state.clear()

    await message.answer(
        f"✅ <b>{data['setting_label']}</b> matni o'zgartirildi!",
        reply_markup=admin_menu_kb(),
    )


# ── Admin: Channels ───────────────────────────────────────────


@router.message(F.text == BTN_ADM_CHANNELS)
async def admin_channels_menu(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    await state.clear()

    channels = await db.get_all_channels()
    if channels:
        text = "📡 <b>Majburiy kanallar:</b>\n\n"
        for i, ch in enumerate(channels, 1):
            title = ch["title"] or ch["channel_id"]
            text += f"  {i}. <b>{title}</b>\n     ID: <code>{ch['channel_id']}</code>\n     🔗 {ch['channel_url']}\n\n"
    else:
        text = "📡 <b>Majburiy kanallar</b>\n\nHozircha kanal qo'shilmagan."

    await message.answer(text, reply_markup=channels_menu_kb())


@router.message(F.text == BTN_CH_LIST)
async def admin_channels_list(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return
    channels = await db.get_all_channels()
    if not channels:
        await message.answer("📡 Kanallar ro'yxati bo'sh.")
        return

    text = "📡 <b>Barcha kanallar:</b>\n\n"
    for i, ch in enumerate(channels, 1):
        title = ch["title"] or ch["channel_id"]
        text += f"  {i}. <b>{title}</b>\n     ID: <code>{ch['channel_id']}</code>\n     🔗 {ch['channel_url']}\n\n"
    await message.answer(text)


@router.message(F.text == BTN_CH_ADD)
async def admin_channel_add_start(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    await message.answer(
        "📡 <b>Kanal qo'shish</b>\n\n"
        "Kanal ID sini kiriting:\n"
        "(Masalan: <code>@polyglotlc</code> yoki <code>-1001234567890</code>)\n\n"
        "⚠️ Bot kanalda admin bo'lishi shart!",
        reply_markup=back_kb(),
    )
    await state.set_state(AdminChannelAdd.waiting_channel_id)


@router.message(AdminChannelAdd.waiting_channel_id)
async def admin_channel_add_id(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    if message.text == BTN_BACK:
        await state.clear()
        await admin_channels_menu(message, state)
        return

    channel_id = message.text.strip()
    await state.update_data(channel_id=channel_id)
    await message.answer(
        "Kanal havolasini kiriting:\n"
        "(Masalan: <code>https://t.me/polyglotlc</code>)",
        reply_markup=back_kb(),
    )
    await state.set_state(AdminChannelAdd.waiting_channel_url)


@router.message(AdminChannelAdd.waiting_channel_url)
async def admin_channel_add_url(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    if message.text == BTN_BACK:
        await message.answer(
            "Kanal ID sini kiriting:",
            reply_markup=back_kb(),
        )
        await state.set_state(AdminChannelAdd.waiting_channel_id)
        return

    data = await state.get_data()
    channel_id = data["channel_id"]
    channel_url = message.text.strip()

    # Try to get channel title
    title = ""
    try:
        chat = await message.bot.get_chat(channel_id)
        title = chat.title or ""
    except Exception:
        pass

    await db.add_channel(channel_id, channel_url, title)
    await state.clear()

    if title:
        await message.answer(
            f"✅ Kanal qo'shildi!\n\n"
            f"📡 <b>{title}</b>\n"
            f"🆔 <code>{channel_id}</code>\n"
            f"🔗 {channel_url}",
            reply_markup=channels_menu_kb(),
        )
    else:
        await message.answer(
            f"⚠️ Kanal saqlandi, lekin ulanishda xato.\n"
            f"Bot kanalda admin ekanligini tekshiring!\n\n"
            f"🆔 <code>{channel_id}</code>\n"
            f"🔗 {channel_url}",
            reply_markup=channels_menu_kb(),
        )


@router.message(F.text == BTN_CH_REMOVE)
async def admin_channel_remove_start(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return

    channels = await db.get_all_channels()
    if not channels:
        await message.answer("📡 O'chirish uchun kanal yo'q.")
        return

    text = "📡 <b>Kanal o'chirish</b>\n\nO'chirmoqchi bo'lgan kanal ID sini kiriting:\n\n"
    for i, ch in enumerate(channels, 1):
        title = ch["title"] or ch["channel_id"]
        text += f"  {i}. <b>{title}</b> — <code>{ch['channel_id']}</code>\n"

    await message.answer(text, reply_markup=back_kb())
    await state.set_state(AdminChannelRemove.waiting_channel_id)


@router.message(AdminChannelRemove.waiting_channel_id)
async def admin_channel_remove_confirm(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    if message.text == BTN_BACK:
        await state.clear()
        await admin_channels_menu(message, state)
        return

    channel_id = message.text.strip()
    removed = await db.remove_channel(channel_id)
    await state.clear()

    if removed:
        await message.answer(
            f"✅ Kanal <code>{channel_id}</code> o'chirildi!",
            reply_markup=channels_menu_kb(),
        )
    else:
        await message.answer(
            f"❌ Kanal <code>{channel_id}</code> topilmadi.",
            reply_markup=channels_menu_kb(),
        )


# ── Admin: Statistics ─────────────────────────────────────────


@router.message(F.text == BTN_ADM_STATS)
async def admin_stats(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    user_count = await db.get_user_count()
    active_count = await db.get_active_user_count()
    top = await db.get_top_referrers(5)
    all_ids = await db.get_all_user_ids()
    channels = await db.get_all_channels()

    text = (
        f"📊 <b>Statistika</b>\n\n"
        f"👥 Ro'yxatdan o'tganlar: <b>{user_count}</b>\n"
        f"✅ Faol (kanalda): <b>{active_count}</b>\n"
        f"📋 Jami foydalanuvchilar: <b>{len(all_ids)}</b>\n"
        f"📡 Kanallar soni: <b>{len(channels)}</b>\n"
    )

    if top:
        text += "\n🏆 <b>Top 5 referalchilar:</b>\n"
        for i, user in enumerate(top, 1):
            name = f"{user['first_name']} {user['last_name']}"
            text += f"  {i}. {name} — {user['ref_count']} ta\n"

    await message.answer(text)


# ── Admin: Users list ─────────────────────────────────────────


@router.message(F.text == BTN_ADM_USERS)
async def admin_users(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import aiosqlite

    async with aiosqlite.connect(db.DB_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute(
            """SELECT u.user_id, u.username, u.first_name, u.last_name, u.phone,
                      u.referred_by, u.registered_at, u.is_registered, u.is_active,
                      (SELECT COUNT(*) FROM users r
                       WHERE r.referred_by = u.user_id AND r.is_registered = 1 AND r.is_active = 1
                      ) as ref_count
               FROM users u
               WHERE u.is_registered = 1
               ORDER BY ref_count DESC, u.registered_at ASC"""
        )
        users = await cursor.fetchall()

    if not users:
        await message.answer("Foydalanuvchilar mavjud emas.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"

    headers = ["#", "Ism", "Familiya", "Telefon", "Username", "Referallar soni", "Holat", "Ro'yxatdan o'tgan"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for i, u in enumerate(users, 1):
        status = "Faol" if u["is_active"] else "Nofaol"
        row_data = [
            i,
            u["first_name"] or "",
            u["last_name"] or "",
            u["phone"] or "",
            f"@{u['username']}" if u["username"] else "",
            u["ref_count"],
            status,
            u["registered_at"] or "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col in (1, 6) else "left", vertical="center")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    file = BufferedInputFile(buffer.read(), filename="foydalanuvchilar.xlsx")
    await message.answer_document(
        file,
        caption=f"📋 <b>Foydalanuvchilar ro'yxati</b>\n\nJami: <b>{len(users)}</b> ta ro'yxatdan o'tgan foydalanuvchi",
    )


# ── Main ──────────────────────────────────────────────────────


async def main():
    if not BOT_TOKEN:
        logger.error("BOT_TOKEN topilmadi! .env faylni tekshiring.")
        return

    await db.init_db()

    bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    dp = Dispatcher()
    dp.include_router(router)

    logger.info("Bot ishga tushdi!")
    await dp.start_polling(bot, allowed_updates=["message", "callback_query", "chat_member"])


if __name__ == "__main__":
    asyncio.run(main())
